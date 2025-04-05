import cv2
import numpy as np
import pandas as pd
from pdf2image import convert_from_bytes
from PIL import Image
import re
import os
import json
import spacy
from azure.ai.vision.imageanalysis import ImageAnalysisClient
from azure.ai.vision.imageanalysis.models import VisualFeatures
from azure.core.credentials import AzureKeyCredential
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from fpdf import FPDF
from io import BytesIO
from form_templates import FORM_TEMPLATES

# Load NLP model
nlp = spacy.load("en_core_web_md")

# Set Azure credentials
AZURE_ENDPOINT = "https://asifali.cognitiveservices.azure.com/"
AZURE_KEY = "EOE9RXRJ4kGTaKDMpkijz3eru1bSarEcRShgxcYEbl9qmpzekr7FJQQJ99BCACGhslBXJ3w3AAAFACOG9wU8"

# Initialize Azure Computer Vision client
vision_client = ImageAnalysisClient(AZURE_ENDPOINT, AzureKeyCredential(AZURE_KEY))

def pdf_to_images(file_bytes):
    """Convert PDF bytes to a list of images."""
    return convert_from_bytes(file_bytes)

def preprocess_image(image):
    """Enhance OCR accuracy with preprocessing."""
    img = np.array(image)
    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
    
    # Adaptive thresholding
    thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 2)
    
    # Denoising & sharpening
    denoised = cv2.fastNlMeansDenoising(thresh, h=15)
    kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]])  # Sharpening kernel
    sharp = cv2.filter2D(denoised, -1, kernel)
    
    return Image.fromarray(sharp)

def extract_text_from_image(image):
    """Extract text using Azure Computer Vision API."""
    processed_image = preprocess_image(image)

    # Convert image to bytes
    image_bytes = BytesIO()
    processed_image.save(image_bytes, format="PNG")
    image_bytes = image_bytes.getvalue()

    # Perform OCR using Azure API
    result = vision_client.analyze(
        image_data=image_bytes,  
        visual_features=[VisualFeatures.Read]  
    )

    extracted_text = []
    if result.read and result.read.blocks:
        for block in result.read.blocks:
            for line in block.lines:
                extracted_text.append(line.text)  

    raw_text = " ".join(extracted_text)
    print("🔍 Raw OCR Output:", raw_text)  # Debugging

    return clean_extracted_text(raw_text) if extracted_text else "No text found"


def clean_extracted_text(text):
    text = re.sub(r"\d:\s+[A-Z]", "", text) 
    text = re.sub(r'\s*\.\s*', ' ', text) 
    text = re.sub(r"[|:{}]", "", text)  # Remove unwanted symbols
    text = re.sub(r"\n{2,}", "\n", text).strip()  # Preserve single newlines
    return text


def extract_fields(text, form_type=None):
    """
    Enhanced extraction of structured fields from text using advanced regex and NLP.
    
    Args:
        text (str): Input text to extract fields from
        form_type (str, optional): Specific form template to use
    
    Returns:
        dict: Structured extracted data
    """
    # Comprehensive data dictionary
    extracted_data = {
        'Personal Information': {},
        'Contact Details': {},
        'Family Information': {},
        'Educational Background': {},
        'Professional Details': {},
        'Miscellaneous': {}
    }
    sections = re.split(r"\n(?=[A-Za-z ]+:)", text)
    for section in sections:
        key_value = section.split(":", 1)
        if len(key_value) == 2:
            key, value = key_value
            extracted_data[key.strip()] = value.strip()
    # Preprocessing
    text = text.replace('\n', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    
    # Personal Information Extraction
    personal_patterns = [
        (r'Full Name[:\s]*([A-Za-z\s]+)(?=Date of Birth|Gender|$)', 'Full Name'),
        (r'Date of Birth[:\s]*(\d{2}/\d{2}/\d{4})', 'Date of Birth'),
        (r'Gender[:\s]*(\w+)', 'Gender')
    ]
    
    for pattern, key in personal_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            extracted_data['Personal Information'][key] = match.group(1).strip()
    
    # Contact Details Extraction
    contact_patterns = [
        (r'Address[:\s]*([^0-9]+)City', 'Address'),
        (r'City[:\s]*([^0-9]+)State', 'City'),
        (r'State[:\s]*(\w+)', 'State'),
        (r'Postal Code[:\s]*(\d+)', 'Postal Code'),
        (r'Phone Number[:\s]*(\d+\s*\d+)', 'Phone Number'),
        (r'Email Address[:\s]*([^\s]+)', 'Email Address')
    ]
    
    for pattern, key in contact_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            extracted_data['Contact Details'][key] = match.group(1).strip()
    
    # Family Information Extraction
    family_patterns = [
        (r"Father's Name[:\s]*([^0-9]+)Father's Occupation", "Father's Name"),
        (r"Father's Occupation[:\s]*([^0-9]+)Mother's Name", "Father's Occupation"),
        (r"Mother's Name[:\s]*([^0-9]+)Mother's Occupation", "Mother's Name"),
        (r"Mother's Occupation[:\s]*([^0-9]+)", "Mother's Occupation")
    ]
    
    for pattern, key in family_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            extracted_data['Family Information'][key] = match.group(1).strip()
    
    # Educational Background Extraction
    edu_patterns = [
        (r'Highest Qualification[:\s]*([^0-9]+)Institution', 'Highest Qualification'),
        (r'Institution Name[:\s]*([^0-9]+)Year', 'Institution Name'),
        (r'Year of Passing[:\s]*(\d{4})', 'Year of Passing'),
        (r'Percentage[:\s]*(\d+)', 'Percentage')
    ]
    
    for pattern, key in edu_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            extracted_data['Educational Background'][key] = match.group(1).strip()
    
    # Professional Details Extraction
    prof_patterns = [
        (r'Current Occupation[:\s]*([^0-9]+)Company', 'Current Occupation'),
        (r'Company Name[:\s]*([^0-9]+)Work', 'Company Name'),
        (r'Work Experience \(years\)[:\s]*(\w+)', 'Work Experience'),
        (r'Key Skills[:\s]*([^0-9]+)Hobbies', 'Key Skills')
    ]
    
    for pattern, key in prof_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            extracted_data['Professional Details'][key] = match.group(1).strip()
    
    # Miscellaneous Extraction
    misc_patterns = [
        (r'Hobbies[:\s]*([^0-9]+)', 'Hobby'),
        (r'Languages Known[:\s]*([^0-9]+)', 'Languages Known')
    ]
    
    for pattern, key in misc_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            extracted_data['Miscellaneous'][key] = match.group(1).strip()
    
    # Remove empty sections
    extracted_data = {k: v for k, v in extracted_data.items() if v}
    
    return extracted_data


def export_to_excel(data_list, output_file="extracted_data.xlsx"):
    """Exports structured data to Excel with formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"
    headers = ["S.No", "Filename", "Extracted Text"]
    ws.append(headers)
    
    for i, item in enumerate(data_list, 1):
        if isinstance(item, tuple) and len(item) == 2:
            filename, text = item
        else:
            filename, text = f"Page {i}", str(item)  # Handle unexpected structures
        ws.append([i, filename, text])
    
    for col_num, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2
    
    wb.save(output_file)

def export_to_pdf(data_list, output_file="extracted_data.pdf"):
    """Exports extracted data to a well-formatted PDF."""
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    
    for item in data_list:
        if isinstance(item, tuple) and len(item) == 2:
            filename, text = item
        else:
            filename, text = "Unknown", str(item)  # Handle unexpected structures

        pdf.cell(200, 10, txt=f"Extracted Data for {filename}", ln=True, align='C')
        pdf.ln(10)
        pdf.multi_cell(0, 10, txt=text)
        pdf.ln(5)
    
    pdf.output(output_file)

def process_pdf(file_path):
    with open(file_path, "rb") as f:
        images = pdf_to_images(f.read())

    extracted_data = []
    for i, img in enumerate(images):
        text = extract_text_from_image(img)
        structured_data = extract_fields(text)

        if text.strip():
            extracted_data.append((f"Page {i+1}", structured_data))

    sorted_data = {k: v for k, v in sorted(extracted_data[0][1].items())}
    return sorted_data

if __name__ == "__main__":
    input_pdf = "sample.pdf"  # Replace with actual PDF file
    extracted_data = process_pdf(input_pdf)
    export_to_excel(extracted_data, "structured_data.xlsx")
    export_to_pdf(extracted_data, "structured_data.pdf")
    print("✅ Extraction Completed with Azure Computer Vision API!") 